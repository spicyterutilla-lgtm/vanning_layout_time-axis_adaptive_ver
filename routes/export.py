from io import BytesIO
import datetime
from flask import jsonify, send_file
from services.session import SESSION_DATA
from services.utils import build_container_alert_summary, build_item_brief
from services.validation import build_data_readiness
from vanning_engine import VOLUME_TARGET_RATE

def export_excel():
    containers = SESSION_DATA.get("last_containers", [])
    if not containers:
        return jsonify({'error': 'エクスポートするデータがありません'}), 400
        
    output = BytesIO()
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # デフォルトシートを削除
    
    # スタイルの定義
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    export_base_date = SESSION_DATA.get("last_base_date", datetime.date.today())
    export_must_ship_window_days = SESSION_DATA.get("last_must_ship_window_days", 7)
    alert_summaries = SESSION_DATA.get("last_alert_summaries", {})
    simulation_context = SESSION_DATA.get("simulation_context") or {}
    is_trial_export = bool(simulation_context.get("is_trial"))
    dataset_prefix = "【試験データ】 " if is_trial_export else ""

    summary_ws = wb.create_sheet(title="納品順サマリー")
    summary_ws.merge_cells('A1:I1')
    summary_ws['A1'] = f"{dataset_prefix}【納品順サマリー】 コンテナ別出荷判断一覧"
    summary_ws['A1'].font = Font(size=14, bold=True)
    summary_headers = [
        "納品順", "計画ID", "最短積込期限", "木箱期限", "体積充填率", "重量充填率",
        "荷物数", "空き埋め荷物", "出荷判断"
    ]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin

    for row_idx, c in enumerate(containers, 4):
        earliest_due = min([item.due_date for item in c.items]) if c.items else None
        exp_items = [item.expiration_date for item in c.items if item.expiration_date]
        earliest_exp = min(exp_items) if exp_items else None
        pull_count = sum(1 for item in c.items if item.status_msg and "前倒し" in item.status_msg)
        alert_summary = alert_summaries.get(c.id) or build_container_alert_summary(c, export_base_date, export_must_ship_window_days)
        decision_text = (
            f"{alert_summary['title']}：{alert_summary['detail']}"
            if alert_summary.get("title")
            else "体積80%クリア"
        )
        values = [
            row_idx - 3,
            c.id,
            earliest_due.strftime("%Y-%m-%d") if earliest_due else "",
            earliest_exp.strftime("%Y-%m-%d") if earliest_exp else "制限なし",
            f"{round(c.fill_rate_volume, 1)}%",
            f"{round(c.fill_rate_weight, 1)}%",
            len(c.items),
            pull_count,
            decision_text,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin
            if col_idx in {1, 7, 8}:
                cell.alignment = Alignment(horizontal="right")
            if c.fill_rate_volume < VOLUME_TARGET_RATE and col_idx == 9:
                cell.font = Font(color="FF0000", bold=True)
            if col_idx == 9:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    summary_widths = [10, 12, 14, 14, 14, 14, 10, 12, 72]
    for col_idx, width in enumerate(summary_widths, 1):
        summary_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    optimization_summary = SESSION_DATA.get("last_optimization_summary")
    if optimization_summary:
        evaluation_ws = wb.create_sheet(title="最適化評価")
        evaluation_ws.merge_cells('A1:C1')
        evaluation_ws['A1'] = "【最適化評価】 算出結果と容量・重量下限"
        evaluation_ws['A1'].font = Font(size=14, bold=True)
        evaluation_rows = [
            ("確定コンテナ本数", optimization_summary.get("container_count", ""), "本"),
            ("容量・重量による理論下限", optimization_summary.get("capacity_lower_bound", ""), "本"),
            ("理論下限との差", optimization_summary.get("container_gap_to_lower_bound", ""), "本"),
            ("体積のみの下限", optimization_summary.get("volume_lower_bound", ""), "本"),
            ("重量のみの下限", optimization_summary.get("weight_lower_bound", ""), "本"),
            ("80%未満コンテナ", optimization_summary.get("alert_containers", ""), "本"),
            ("体積総量上避けられない80%未満", optimization_summary.get("volume_only_minimum_alerts", ""), "本"),
            ("平均体積充填率", optimization_summary.get("avg_volume_rate", ""), "%"),
            ("平均重量充填率", optimization_summary.get("avg_weight_rate", ""), "%"),
            ("比較した初期配置候補", optimization_summary.get("strategy_trial_count", ""), "案"),
            ("最終再配置まで比較した候補", optimization_summary.get("deep_strategy_trial_count", ""), "案"),
            ("採用した配置方針", optimization_summary.get("selected_strategy", ""), ""),
            ("3D配置チェック", "適合" if optimization_summary.get("geometry_valid") else "要確認", ""),
            ("3D配置警告", optimization_summary.get("geometry_warning_count", 0), "件"),
        ]
        for row_idx, (label, value, unit) in enumerate(evaluation_rows, 3):
            evaluation_ws.cell(row=row_idx, column=1).value = label
            evaluation_ws.cell(row=row_idx, column=2).value = value
            evaluation_ws.cell(row=row_idx, column=3).value = unit
            for col_idx in range(1, 4):
                cell = evaluation_ws.cell(row=row_idx, column=col_idx)
                cell.border = border_thin
                if col_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        note_row = len(evaluation_rows) + 5
        evaluation_ws.cell(row=note_row, column=1).value = "注記"
        evaluation_ws.cell(row=note_row, column=1).fill = header_fill
        evaluation_ws.cell(row=note_row, column=1).font = header_font
        evaluation_ws.cell(row=note_row, column=1).border = border_thin
        evaluation_ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=3)
        evaluation_ws.cell(row=note_row, column=2).value = optimization_summary.get("lower_bound_note", "")
        evaluation_ws.cell(row=note_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        evaluation_ws.cell(row=note_row, column=2).border = border_thin
        evaluation_ws.column_dimensions["A"].width = 34
        evaluation_ws.column_dimensions["B"].width = 70
        evaluation_ws.column_dimensions["C"].width = 10

    if simulation_context:
        def pct(value):
            return "" if value is None else f"{value}%"

        context_ws = wb.create_sheet(title="前提サマリー")
        context_ws.merge_cells('A1:D1')
        context_ws['A1'] = "【前提サマリー】 今回のデータの根拠と仮定"
        context_ws['A1'].font = Font(size=14, bold=True)
        context_ws['A2'] = simulation_context.get("summary_note", "")
        context_ws.merge_cells('A2:D2')
        context_ws['A2'].alignment = Alignment(wrap_text=True, vertical="top")

        data_type = (
            "試験データ（別枠読込）"
            if simulation_context.get("is_trial")
            else "開示ケースリスト準拠の検証データ"
            if simulation_context.get("is_generated")
            else "読み込みExcelデータ"
        )
        context_rows = [
            ("データ種別", data_type),
            ("入力ファイル", simulation_context.get("source_filename", "")),
            ("前提プロファイル", simulation_context.get("assumption_profile_name", "")),
            ("前提バージョン", simulation_context.get("assumption_profile_version", "")),
            ("非開示値の扱い", simulation_context.get("assumption_policy", "")),
            ("読み込み荷物数", simulation_context.get("loaded_count", "")),
            ("生成時の荷物数", simulation_context.get("generated_rows_setting", "")),
            ("再現用番号", simulation_context.get("seed", "")),
            ("出荷情報受領日", simulation_context.get("cargo_information_date", "")),
            ("レイアウト確定期限", simulation_context.get("layout_confirmation_date", "")),
            ("バンニング期間", f"{simulation_context.get('vanning_start_date', '')} 〜 {simulation_context.get('vanning_end_date', '')}"),
            ("船積日", simulation_context.get("vessel_loading_date", "")),
            ("積み方", simulation_context.get("placement_policy", "")),
            ("木箱期限あり", pct(simulation_context.get("wood_deadline_rate"))),
            ("納入状況", f"{simulation_context.get('availability_reference_label', '')}: 納入予定済み {pct(simulation_context.get('arrived_rate'))} / 期限後予定 {pct(simulation_context.get('future_rate'))}"),
            ("バンニング完了期限範囲", f"{simulation_context.get('due_min', '')} 〜 {simulation_context.get('due_max', '')}"),
            ("平均重量", simulation_context.get("avg_weight", "")),
        ]
        for row_idx, (label, value) in enumerate(context_rows, 4):
            context_ws.cell(row=row_idx, column=1).value = label
            context_ws.cell(row=row_idx, column=2).value = value
            for col_idx in (1, 2):
                cell = context_ws.cell(row=row_idx, column=col_idx)
                cell.border = border_thin
                if col_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        source_start = 25
        source_headers = ["項目", "扱い", "根拠", "備考"]
        for col_idx, header in enumerate(source_headers, 1):
            cell = context_ws.cell(row=source_start, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin
        for row_idx, row in enumerate(simulation_context.get("source_rows", []), source_start + 1):
            values = [row.get("item", ""), row.get("status", ""), row.get("basis", ""), row.get("note", "")]
            for col_idx, value in enumerate(values, 1):
                cell = context_ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border_thin
                if row.get("status") in {"仮定", "採用仮定", "補完"} and col_idx == 2:
                    cell.font = Font(color="B45309", bold=True)
                if col_idx == 4:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, width in enumerate([24, 18, 24, 72], 1):
            context_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    readiness = build_data_readiness(SESSION_DATA.get("input_profile", {}), SESSION_DATA.get("validation_issues", []))
    readiness_ws = wb.create_sheet(title="実務準備度")
    readiness_ws.merge_cells('A1:E1')
    readiness_ws['A1'] = f"【実務準備度】 {readiness['score']}% / {readiness['risk_level']}"
    readiness_ws['A1'].font = Font(size=14, bold=True)
    readiness_headers = ["項目", "根拠区分", "確認列", "備考", "入力列"]
    for col_idx, header in enumerate(readiness_headers, 1):
        cell = readiness_ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin
    for row_idx, row in enumerate(readiness["checks"], 4):
        values = [row["item"], row["status"], row["source"], row["note"], "あり" if row["ready"] else "補完"]
        for col_idx, value in enumerate(values, 1):
            cell = readiness_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin
            if row["status"] == "採用仮定" and col_idx == 2:
                cell.font = Font(color="B45309", bold=True)
            if not row["ready"] and col_idx in {2, 5}:
                cell.font = Font(color="FF0000", bold=True)
            if col_idx == 4:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, width in enumerate([18, 14, 24, 72, 12], 1):
        readiness_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    for c in containers:
        ws = wb.create_sheet(title=str(c.id))
        alert_summary = alert_summaries.get(c.id) or build_container_alert_summary(c, export_base_date, export_must_ship_window_days)
        alert_text = (
            f"{alert_summary['title']}：{alert_summary['detail']}"
            if alert_summary["title"]
            else "体積80%クリア"
        )
        
        # コンテナサマリー部分
        ws.merge_cells('A1:M1')
        ws['A1'] = f"{dataset_prefix}【バンニング指示書】 コンテナ: {c.id}"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = "最大重量(kg)"
        ws['B3'] = c.max_weight
        ws['A4'] = "現在重量(kg)"
        ws['B4'] = c.current_weight
        ws['A5'] = "重量充填率"
        ws['B5'] = f"{round(c.fill_rate_weight, 1)}%"
        ws['A6'] = "体積充填率"
        ws['B6'] = f"{round(c.fill_rate_volume, 1)}%"
        ws['A7'] = "出荷判断"
        ws['B7'] = alert_text
        ws.merge_cells('B7:M7')
        
        for r in range(3, 8):
            ws[f'A{r}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws[f'A{r}'].border = border_thin
            ws[f'B{r}'].border = border_thin
        ws['B7'].alignment = Alignment(wrap_text=True, vertical="top")
        if c.fill_rate_volume < VOLUME_TARGET_RATE:
            ws['B7'].font = Font(color="FF0000", bold=True)
        
        # 荷物リストのヘッダー
        headers = [
            "積込順", "依頼コード", "部品コード", "車両コード", "仕向け地", "品名",
            "梱包前重量(kg)", "梱包後重量(kg)", "寸法 L×W×H(mm)",
            "3D配置 (奥×横×高)", "分類理由", "判断根拠", "特記事項・ステータス"
        ]
        start_row = 10
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin
            
        # 荷物データの書き込み
        for idx, i in enumerate(c.items, 1):
            row_idx = start_row + idx
            
            # 回転考慮の寸法
            dim_str = f"{i.length} × {i.width} × {i.height}" + (" (回転)" if i.is_rotated else "")
            
            # 3D座標 (Three.jsの表示に合わせて見やすく)
            pos_str = f"X:{i.x} Y:{i.y} Z:{i.z}" if i.x is not None else "未定"
            
            # 現場ステータス
            status = i.status_msg if i.status_msg else "通常"
            if i.is_force_ship and status == "通常":
                status = "80%未満/積込期限優先出荷"
                
            row_data = [
                idx,
                i.request_code,
                i.part_code,
                i.vehicle_code,
                i.destination,
                i.name,
                f"{i.unpacked_weight:,.1f}" if i.unpacked_weight is not None else "",
                f"{i.weight:,.1f}",
                dim_str,
                pos_str,
                i.selection_reason if i.selection_reason else "通常",
                i.decision_reason or i.selection_reason,
                status
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.border = border_thin
                if col_idx in {1, 7, 8}:
                    cell.alignment = Alignment(horizontal="right")
                    
                # 色付けルールの適用
                if col_idx == 13:
                    if "強制出荷" in status:
                        cell.font = Font(color="FF0000", bold=True) # 赤
                    elif "前倒し" in status:
                        cell.font = Font(color="008000", bold=True) # 緑
        
        # 列幅の自動調整
        widths = [8, 22, 16, 14, 20, 30, 15, 15, 25, 25, 20, 48, 40]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    review_queue = SESSION_DATA.get("last_review_queue", [])
    if review_queue:
        ws = wb.create_sheet(title="ReviewQueue")
        headers = [
            "品名", "確認理由", "納入予定日", "積込期限", "木箱期限", "積込期限まで(日)",
            "木箱期限まで(日)", "優先度", "分類理由", "判断根拠", "ステータス"
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin

        for row_idx, item in enumerate(review_queue, 2):
            row_data = [
                item["name"],
                item["reason"],
                item["creation_date"],
                item["due_date"],
                item["expiration_date"],
                item["days_until_due"],
                item["days_until_expiration"],
                item["priority"],
                item["selection_reason"],
                item["decision_reason"],
                item["status_msg"],
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.border = border_thin

        widths = [30, 36, 14, 14, 14, 14, 16, 10, 20, 48, 30]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(output)
    output.seek(0)
    return send_file(
        output, 
        download_name='trial_vanning_instructions.xlsx' if is_trial_export else 'vanning_instructions.xlsx',
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
