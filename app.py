import os
import datetime
import uuid
import copy
import json
import pandas as pd
from io import BytesIO
from flask import Flask, request, jsonify, send_from_directory, send_file
from data_loader import DataLoader
from models import Container
from simulation_assumptions import build_generation_parameter_rows, load_simulation_assumptions
from vanning_engine import VanningEngine, VOLUME_TARGET_RATE

app = Flask(__name__, static_folder='static')
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# 簡易的なオンメモリセッション（プロトタイプ用）
SESSION_DATA = {"items": [], "validation_issues": [], "invalid_item_ids": set()}
SYSTEM_STATUS_KEYWORDS = ("前倒し", "保留プール", "赤字強制出荷")
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
ROLLING_PLANNING_WINDOW_DAYS = 7
ROLLING_FORCE_SHIP_WINDOW_DAYS = 0
ROLLING_PLANNING_WINDOW_CANDIDATES = (0, 3, 5, 7, 10, 14, 21)
CASE_MASTER_FILE = "(抜粋)ケースリスト.xlsx"
SIMULATION_ASSUMPTIONS_FILE = "simulation_assumptions.json"
CASE_MASTER_CACHE = {"path": None, "mtime": None, "rows": None}

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('static', path)

@app.route('/api/status', methods=['GET'])
def get_status():
    items = SESSION_DATA.get("items", [])
    if items:
        validation_issues = SESSION_DATA.get("validation_issues", [])
        return jsonify({
            'has_data': True,
            'total_items': len(items),
            'validation_summary': build_validation_summary(validation_issues),
            'validation_issues': validation_issues[:8],
            'readiness': build_data_readiness(SESSION_DATA.get("input_profile", {}), validation_issues),
            'simulation_context': SESSION_DATA.get("simulation_context")
        })
    return jsonify({'has_data': False})

def can_fit_container(item, container=None):
    container = container or Container(id="VALIDATION")
    if item.height > container.height or item.weight > container.max_weight:
        return False
    normal = item.length <= container.length and item.width <= container.width
    rotated = item.width <= container.length and item.length <= container.width
    return normal or rotated

def validate_items(items):
    issues = []
    invalid_item_ids = set()
    validation_container = Container(id="VALIDATION")

    for item in items:
        item_issues = []

        if item.length <= 0 or item.width <= 0 or item.height <= 0:
            item_issues.append(("error", "寸法が0以下です"))
        if item.weight <= 0:
            item_issues.append(("error", "重量が0以下です"))
        if item.due_date < item.creation_date:
            item_issues.append(("error", "納期が積載可能日より前です"))
        if item.expiration_date and item.expiration_date < item.creation_date:
            item_issues.append(("error", "保管期限が積載可能日より前です"))
        if not can_fit_container(item, validation_container):
            item_issues.append(("error", "有効内寸または最大重量を超えており積載できません"))

        if item.expiration_date and item.expiration_date < item.due_date:
            item_issues.append(("warning", "木箱期限が納期より前です"))

        for severity, message in item_issues:
            issues.append({
                "severity": severity,
                "item_id": item.id,
                "name": item.name,
                "message": message
            })
            if severity == "error":
                invalid_item_ids.add(item.id)

    return issues, invalid_item_ids

def build_validation_summary(issues):
    return {
        "errors": sum(1 for issue in issues if issue["severity"] == "error"),
        "warnings": sum(1 for issue in issues if issue["severity"] == "warning"),
        "total": len(issues)
    }

def build_data_readiness(input_profile, validation_issues):
    detected = (input_profile or {}).get("detected_columns", {})
    is_generated_simulation = bool((input_profile or {}).get("generated_simulation"))
    checks = [
        ("ケース寸法", True, "L/W/H/名称", "開示ケースマスタまたは入力Excelから取得"),
        ("重量", detected.get("weight", False), "重量列", "未入力の場合は1000kg補完または生成仮定になり、実務検証では要確認"),
        ("積載可能日", detected.get("creation", False), "積載可能日/到着日", "未入力の場合は当日扱いになり、時間軸評価が粗くなる"),
        ("納期", detected.get("due", False), "納期", "未入力の場合は7日後扱いになり、出荷判断の根拠が弱くなる"),
        ("木箱期限", detected.get("expiration", False), "木箱期限", "未入力の場合は木箱のみ21日補完または生成仮定"),
        ("前倒し可否", detected.get("allow_early_ship", False), "前倒し可否", "未入力の場合は前倒し可として扱う"),
        ("段積み制約", detected.get("stackable", False), "段積み可否/段積み不可", "未入力の場合は段積み可として扱う"),
        ("床置き制約", detected.get("floor_only", False), "床置き", "未入力の場合は床置き指定なしとして扱う"),
        ("分離制約", detected.get("separation_group", False), "分離グループ", "未入力の場合は混載制約なしとして扱う"),
    ]
    rows = []
    ready_count = 0
    for item, ready, source, note in checks:
        if ready:
            ready_count += 1
        rows.append({
            "item": item,
            "status": "実データあり" if ready else "仮定/補完",
            "ready": ready,
            "source": source,
            "note": note,
        })

    score = round((ready_count / len(checks)) * 100, 1) if checks else 0
    if is_generated_simulation:
        score = min(score, 55.0)
        rows.insert(0, {
            "item": "データ性質",
            "status": "仮定シミュレーション",
            "ready": False,
            "source": "仮定根拠/仮定重量階級",
            "note": "開示ケース寸法を根拠にした検証用データ。重量・頻度・納期・制約は先方回答で差し替え対象",
        })
    risk_level = "実務検証向き" if score >= 80 else "要追加確認" if score >= 55 else "仮定検証向き"
    return {
        "score": score,
        "risk_level": risk_level,
        "checks": rows,
        "validation": build_validation_summary(validation_issues),
    }

def read_generation_parameters(filepath):
    try:
        df = pd.read_excel(filepath, sheet_name="生成パラメータ")
    except Exception:
        return {}

    params = {}
    for _, row in df.iterrows():
        key = str(row.get("項目", "")).strip()
        if key:
            params[key] = row.get("値", "")
    return params

def coerce_float(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return float(str(value).replace("%", "").strip())
    except (TypeError, ValueError):
        return None

def parameter_rate_to_percent(value):
    numeric = coerce_float(value)
    if numeric is None:
        return None
    return round(numeric * 100, 1) if 0 <= numeric <= 1 else round(numeric, 1)

def safe_rate(count, total):
    return round((count / total) * 100, 1) if total else 0

def build_simulation_context(items, input_profile=None, generation_parameters=None):
    if not items:
        return None

    input_profile = input_profile or {}
    generation_parameters = generation_parameters or {}
    total = len(items)
    today = datetime.date.today()
    due_dates = [item.due_date for item in items if item.due_date]
    weights = [item.weight for item in items if item.weight is not None]

    early_count = sum(1 for item in items if item.allow_early_ship)
    separation_count = sum(1 for item in items if item.separation_group)
    no_stack_count = sum(1 for item in items if not item.stackable)
    floor_only_count = sum(1 for item in items if item.floor_only)
    wood_deadline_count = sum(1 for item in items if item.expiration_date)
    future_count = sum(1 for item in items if item.creation_date > today)
    arrived_count = total - future_count
    is_generated = bool(input_profile.get("generated_simulation"))
    detected = input_profile.get("detected_columns", {})
    if is_generated:
        source_rows = [
            {"item": "ケース寸法・箱タイプ", "status": "開示資料ベース", "basis": "ケースマスタ", "note": "いすゞロジスティック開示ケースリストの寸法を使用"},
            {"item": "重量", "status": "仮定", "basis": "重量密度レンジ", "note": "先方回答が来たら実績重量・階級で差し替え"},
            {"item": "出現頻度", "status": "仮定", "basis": "ケース階級比率", "note": "実際の週次・月次出現比率で更新対象"},
            {"item": "積載可能日・納期", "status": "仮定", "basis": "日付分布", "note": "入荷日・納期分布の回答待ち"},
            {"item": "木箱期限", "status": "仮定", "basis": "木箱期限候補", "note": "木箱保管期限ルールの回答待ち"},
            {"item": "段積み・床置き・分離制約", "status": "仮定", "basis": "制約発生率", "note": "現場制約ルールの回答待ち"},
        ]
        summary_note = "寸法は開示資料ベース、重量・頻度・日付・制約は暫定仮定です。"
    else:
        source_rows = [
            {"item": "ケース寸法・品名", "status": "入力Excel", "basis": "L/W/H/名称", "note": "読み込んだExcelの値を使用"},
            {"item": "重量", "status": "入力Excel" if detected.get("weight") else "補完", "basis": "重量列", "note": "列がない場合は補完値のため要確認"},
            {"item": "積載可能日", "status": "入力Excel" if detected.get("creation") else "補完", "basis": "積載可能日/到着日", "note": "列がない場合は当日扱い"},
            {"item": "納期", "status": "入力Excel" if detected.get("due") else "補完", "basis": "納期", "note": "列がない場合は7日後扱い"},
            {"item": "木箱期限", "status": "入力Excel" if detected.get("expiration") else "補完", "basis": "木箱期限", "note": "列がない場合は木箱のみ補完"},
            {"item": "現場制約", "status": "入力Excel" if any(detected.get(key) for key in ("allow_early_ship", "stackable", "floor_only", "separation_group")) else "補完", "basis": "前倒し/段積み/床置き/分離", "note": "不足列は既定値で補完"},
        ]
        summary_note = "読み込んだExcelの列を優先し、不足項目は安全側の既定値で補完しています。"

    return {
        "is_generated": is_generated,
        "loaded_count": total,
        "generated_rows_setting": coerce_float(generation_parameters.get("生成行数")),
        "seed": str(generation_parameters.get("乱数シード", "")).strip(),
        "configured_early_ship_rate": parameter_rate_to_percent(generation_parameters.get("前倒し可率")),
        "configured_separation_rate": parameter_rate_to_percent(generation_parameters.get("分離制約発生率")),
        "early_ship_rate": safe_rate(early_count, total),
        "separation_rate": safe_rate(separation_count, total),
        "no_stack_rate": safe_rate(no_stack_count, total),
        "floor_only_rate": safe_rate(floor_only_count, total),
        "wood_deadline_rate": safe_rate(wood_deadline_count, total),
        "arrived_rate": safe_rate(arrived_count, total),
        "future_rate": safe_rate(future_count, total),
        "due_min": min(due_dates).strftime("%Y-%m-%d") if due_dates else "",
        "due_max": max(due_dates).strftime("%Y-%m-%d") if due_dates else "",
        "avg_weight": round(sum(weights) / len(weights)) if weights else 0,
        "summary_note": summary_note,
        "source_rows": source_rows,
    }

def clear_last_results():
    for key in ("last_containers", "last_rolling", "last_review_queue", "last_scenarios"):
        SESSION_DATA.pop(key, None)

def reset_runtime_state(items):
    """再計算前に、前回の配置結果とシステム判定だけをクリアする。"""
    for item in items:
        item.x = item.y = item.z = None
        item.is_rotated = False
        item.selection_reason = ""
        item.decision_reason = ""
        if hasattr(item, "system_force_ship"):
            item.system_force_ship = False

        if any(keyword in str(item.status_msg) for keyword in SYSTEM_STATUS_KEYWORDS):
            item.status_msg = ""

        if item.force_ship and not item.status_msg:
            item.status_msg = "手動オーバーライド：現場指示による強制出荷"

def classify_items_for_day(items, current_date, must_ship_window_days=7):
    """
    plan_date 時点で荷物を Must / Forwardable / Future に分類する。
    Future は未到着なので、当日の3Dパッキングには渡さない。
    """
    target_date_limit = current_date + datetime.timedelta(days=must_ship_window_days)
    groups = {
        "must_ship": [],
        "forwardable": [],
        "hold": [],
        "future": [],
    }

    for item in items:
        if item.creation_date > current_date:
            item.selection_reason = "未到着"
            groups["future"].append(item)
            continue

        must_reasons = []
        if item.force_ship:
            must_reasons.append("手動強制")
        if item.due_date <= target_date_limit:
            must_reasons.append("納期接近")
        if item.expiration_date and item.expiration_date <= target_date_limit:
            must_reasons.append("木箱期限接近")

        if must_reasons:
            item.selection_reason = " / ".join(must_reasons)
            groups["must_ship"].append(item)
        elif item.allow_early_ship:
            item.selection_reason = "前倒し候補"
            groups["forwardable"].append(item)
        else:
            item.selection_reason = "前倒し不可"
            groups["hold"].append(item)

    return groups

def build_upload_path(original_filename):
    _, ext = os.path.splitext(original_filename or "")
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError("Excelファイル（.xlsx / .xls）のみアップロードできます")
    return os.path.join(app.config['UPLOAD_FOLDER'], f"{uuid.uuid4().hex}{ext}")

def summarize_containers(containers):
    if not containers:
        return {
            "container_count": 0,
            "alert_containers": 0,
            "avg_volume_rate": 0,
            "avg_weight_rate": 0,
        }

    return {
        "container_count": len(containers),
        "alert_containers": sum(1 for c in containers if c.fill_rate_volume < VOLUME_TARGET_RATE),
        "avg_volume_rate": round(sum(c.fill_rate_volume for c in containers) / len(containers), 1),
        "avg_weight_rate": round(sum(c.fill_rate_weight for c in containers) / len(containers), 1),
    }

def output_container_sort_key(container):
    earliest_due = None
    earliest_expiration = None
    for item in container.items:
        earliest_due = item.due_date if earliest_due is None else min(earliest_due, item.due_date)
        if item.expiration_date:
            earliest_expiration = (
                item.expiration_date
                if earliest_expiration is None
                else min(earliest_expiration, item.expiration_date)
            )
    return (
        earliest_due or datetime.date.max,
        earliest_expiration or datetime.date.max,
        container.id,
    )

def order_containers_for_output(containers):
    return sorted(containers, key=output_container_sort_key)

def build_comparison(baseline_containers, optimized_containers):
    baseline = summarize_containers(baseline_containers)
    optimized = summarize_containers(optimized_containers)
    return {
        "baseline": baseline,
        "optimized": optimized,
        "delta": {
            "container_count": baseline["container_count"] - optimized["container_count"],
            "alert_containers": baseline["alert_containers"] - optimized["alert_containers"],
            "avg_volume_rate": round(optimized["avg_volume_rate"] - baseline["avg_volume_rate"], 1),
            "avg_weight_rate": round(optimized["avg_weight_rate"] - baseline["avg_weight_rate"], 1),
        }
    }

def build_valid_items():
    items = SESSION_DATA.get("items", [])
    invalid_item_ids = SESSION_DATA.get("invalid_item_ids", set())
    return [item for item in items if item.id not in invalid_item_ids], invalid_item_ids

def read_bounded_int(data, key, default, min_value, max_value):
    try:
        value = int(data.get(key, default))
    except (TypeError, ValueError):
        value = default
    return max(min_value, min(value, max_value))

def parse_base_date(value):
    if not value:
        return datetime.date.today(), None
    try:
        return datetime.datetime.strptime(value, '%Y-%m-%d').date(), None
    except (TypeError, ValueError):
        return None, "作業基準日は YYYY-MM-DD 形式で指定してください"

def nearest_deadline(item):
    candidates = [item.due_date]
    if item.expiration_date:
        candidates.append(item.expiration_date)
    return min(candidates)

def build_constraint_tags(item):
    tags = []
    if not item.stackable:
        tags.append("段積み不可")
    if item.floor_only:
        tags.append("床置き")
    if item.separation_group:
        tags.append(f"分離:{item.separation_group}")
    return tags

def item_bounds(item):
    length = item.width if item.is_rotated else item.length
    width = item.length if item.is_rotated else item.width
    x = item.x if item.x is not None else 0
    y = item.y if item.y is not None else 0
    z = item.z if item.z is not None else 0
    return x, y, z, x + length, y + width, z + item.height

def ranges_overlap(a_start, a_end, b_start, b_end):
    return a_start < b_end and b_start < a_end

def boxes_overlap(a, b):
    ax1, ay1, az1, ax2, ay2, az2 = item_bounds(a)
    bx1, by1, bz1, bx2, by2, bz2 = item_bounds(b)
    return (
        ranges_overlap(ax1, ax2, bx1, bx2)
        and ranges_overlap(ay1, ay2, by1, by2)
        and ranges_overlap(az1, az2, bz1, bz2)
    )

def footprint_contains(support, item):
    sx1, sy1, _, sx2, sy2, sz2 = item_bounds(support)
    ix1, iy1, iz1, ix2, iy2, _ = item_bounds(item)
    return (
        abs(sz2 - iz1) < 0.001
        and sx1 <= ix1
        and sy1 <= iy1
        and sx2 >= ix2
        and sy2 >= iy2
    )

def validate_container_geometry(container):
    warnings = []
    for item in container.items:
        x1, y1, z1, x2, y2, z2 = item_bounds(item)
        if x1 < -0.001 or y1 < -0.001 or z1 < -0.001 or x2 > container.length + 0.001 or y2 > container.width + 0.001 or z2 > container.height + 0.001:
            warnings.append(f"{item.name}: コンテナ内寸外")
        if item.floor_only and z1 > 0.001:
            warnings.append(f"{item.name}: 床置き指定違反")
        if z1 > 0.001 and not any(footprint_contains(other, item) for other in container.items if other is not item):
            warnings.append(f"{item.name}: 底面支持なし")

    for idx, item in enumerate(container.items):
        for other in container.items[idx + 1:]:
            if boxes_overlap(item, other):
                warnings.append(f"{item.name} / {other.name}: 3D重なり")

    return {
        "valid": not warnings,
        "warnings": warnings[:10],
        "warning_count": len(warnings),
    }

def build_forwardable_blocker_counts(engine, container, candidate_items):
    counts = {}
    if container.fill_rate_volume >= VOLUME_TARGET_RATE:
        return counts
    for item in candidate_items:
        reason = engine.explain_unfit_reason(container, item)
        counts[reason] = counts.get(reason, 0) + 1
    return counts

def format_blocker_counts(blocker_counts):
    if not blocker_counts:
        return ""
    order = ["重量上限", "重量ソフト上限", "3D配置/床置き/段積み", "分離制約", "採用可能"]
    parts = []
    for key in order:
        if blocker_counts.get(key):
            parts.append(f"{key}{blocker_counts[key]}件")
    for key, value in blocker_counts.items():
        if key not in order and value:
            parts.append(f"{key}{value}件")
    return " / ".join(parts[:3])

def build_container_alert_summary(container, current_date, must_ship_window_days, blocker_counts=None):
    is_alert = container.fill_rate_volume < VOLUME_TARGET_RATE
    pulled_count = sum(1 for item in container.items if item.status_msg and "前倒し" in item.status_msg)
    if not is_alert:
        if pulled_count:
            return {
                "title": "前倒し補填で80%クリア",
                "detail": f"前倒し補填{pulled_count}件を採用し、体積充填率{container.fill_rate_volume:.1f}%まで改善しました。"
            }
        return {"title": "", "detail": ""}

    limit_date = current_date + datetime.timedelta(days=must_ship_window_days)
    manual_count = sum(1 for item in container.items if item.force_ship)
    due_count = sum(1 for item in container.items if item.due_date <= limit_date)
    exp_count = sum(1 for item in container.items if item.expiration_date and item.expiration_date <= limit_date)

    reasons = []
    if manual_count:
        reasons.append(f"手動指定{manual_count}件")
    if due_count:
        reasons.append(f"納期{must_ship_window_days}日以内{due_count}件")
    if exp_count:
        reasons.append(f"木箱期限{must_ship_window_days}日以内{exp_count}件")
    if not reasons:
        reasons.append("期限優先の必須出荷")

    supplement = (
        f"前倒し補填{pulled_count}件後も体積{container.fill_rate_volume:.1f}%です。"
        if pulled_count
        else "前倒し候補は3D配置・重量・実務制約により追加採用できませんでした。"
    )
    blocker_text = format_blocker_counts(blocker_counts or {})
    if blocker_text:
        supplement += f" 未採用候補の主因: {blocker_text}。"
    return {
        "title": "80%未満でも期限優先で出荷",
        "detail": f"{' / '.join(reasons)}のため保留不可。{supplement}"
    }

def build_item_brief(item, reference_date, reason):
    days_until_due = (item.due_date - reference_date).days
    days_until_expiration = (item.expiration_date - reference_date).days if item.expiration_date else None
    return {
        "id": item.id,
        "name": item.name,
        "reason": reason,
        "creation_date": item.creation_date.strftime("%Y-%m-%d"),
        "due_date": item.due_date.strftime("%Y-%m-%d"),
        "expiration_date": item.expiration_date.strftime("%Y-%m-%d") if item.expiration_date else "",
        "days_until_due": days_until_due,
        "days_until_expiration": days_until_expiration,
        "priority": item.priority,
        "allow_early_ship": item.allow_early_ship,
        "constraint_tags": build_constraint_tags(item),
        "selection_reason": item.selection_reason,
        "decision_reason": item.decision_reason or item.selection_reason or reason,
        "status_msg": item.status_msg,
    }

def review_sort_key(entry):
    exp_days = entry["days_until_expiration"]
    nearest_days = min(
        entry["days_until_due"],
        exp_days if exp_days is not None else entry["days_until_due"]
    )
    risk_rank = 0 if nearest_days < 0 else 1 if nearest_days <= 7 else 2
    return (risk_rank, nearest_days, -entry["priority"], entry["due_date"], entry["name"])

def build_review_queue(pool_items, unused_forwardable, hold_items, future_items, current_date, limit=12):
    seen = set()
    entries = []
    sources = [
        (pool_items, "保留プール: 3D配置または充填率判断で今回未確定"),
        (unused_forwardable, "未使用の前倒し候補: 今回の赤字補填には未採用"),
        (hold_items, "前倒し不可: 納期または木箱期限まで待機"),
        (future_items, "未到着: 積載可能日待ち"),
    ]

    for items, reason in sources:
        for item in items:
            if item.id in seen:
                continue
            seen.add(item.id)
            entries.append(build_item_brief(item, current_date, reason))

    return sorted(entries, key=review_sort_key)[:limit]

def build_remaining_risks(remaining_items, reference_date, limit=20):
    entries = []
    for item in remaining_items:
        if item.due_date <= reference_date or (item.expiration_date and item.expiration_date <= reference_date):
            reason = "期間終了時点で期限到来/超過"
        elif item.creation_date > reference_date:
            reason = "期間内に未到着"
        elif not item.allow_early_ship:
            reason = "前倒し不可で待機"
        else:
            reason = "期間内で未出荷"
        entries.append(build_item_brief(item, reference_date, reason))

    return sorted(entries, key=review_sort_key)[:limit]

def run_rolling_simulation(items, start_date, days=30, allow_pull=True, planning_window_days=0, force_ship_window_days=0):
    ledger_items = copy.deepcopy(items)
    shipped_ids = set()
    daily_results = []
    total_containers = 0
    total_alert_containers = 0
    total_pulls = 0
    volume_rates = []
    weight_rates = []

    for offset in range(days):
        plan_date = start_date + datetime.timedelta(days=offset)
        remaining_items = [item for item in ledger_items if item.id not in shipped_ids]
        reset_runtime_state(remaining_items)
        day_groups = classify_items_for_day(remaining_items, plan_date, must_ship_window_days=planning_window_days)

        engine = VanningEngine()
        forwardable_items = day_groups["forwardable"] if allow_pull else []
        containers, pool, unused_forwardable = engine.run_time_axis_packing(
            day_groups["must_ship"],
            forwardable_items,
            plan_date,
            force_ship_window_days=force_ship_window_days,
            allow_partial_red_pull=allow_pull
        )

        shipped_today_ids = {item.id for c in containers for item in c.items}
        shipped_ids.update(shipped_today_ids)

        shipped_today = sum(len(c.items) for c in containers)
        pull_count = sum(
            1
            for c in containers
            for item in c.items
            if item.status_msg and "前倒し" in item.status_msg
        )
        alerts = sum(1 for c in containers if c.fill_rate_volume < VOLUME_TARGET_RATE)
        avg_volume = round(sum(c.fill_rate_volume for c in containers) / len(containers), 1) if containers else 0
        avg_weight = round(sum(c.fill_rate_weight for c in containers) / len(containers), 1) if containers else 0

        total_containers += len(containers)
        total_alert_containers += alerts
        total_pulls += pull_count
        if containers:
            volume_rates.extend(c.fill_rate_volume for c in containers)
            weight_rates.extend(c.fill_rate_weight for c in containers)

        deadline_remaining = sum(
            1 for item in remaining_items
            if item.id not in shipped_today_ids
            and (item.due_date <= plan_date or (item.expiration_date and item.expiration_date <= plan_date))
        )

        daily_results.append({
            "date": plan_date.strftime("%Y-%m-%d"),
            "containers": len(containers),
            "shipped": shipped_today,
            "pulls": pull_count,
            "alerts": alerts,
            "must_ship": len(day_groups["must_ship"]),
            "forwardable": len(day_groups["forwardable"]),
            "unused_forwardable": len(unused_forwardable) if allow_pull else len(day_groups["forwardable"]),
            "hold": len(day_groups["hold"]),
            "future": len(day_groups["future"]),
            "pool": len(pool),
            "overdue_remaining": deadline_remaining,
            "deadline_remaining": deadline_remaining,
            "avg_volume_rate": avg_volume,
            "avg_weight_rate": avg_weight,
        })

    remaining_items = [item for item in ledger_items if item.id not in shipped_ids]
    period_end = start_date + datetime.timedelta(days=days - 1)
    overdue_count = sum(
        1 for item in remaining_items
        if item.due_date <= period_end or (item.expiration_date and item.expiration_date <= period_end)
    )

    return {
        "start_date": start_date.strftime("%Y-%m-%d"),
        "end_date": period_end.strftime("%Y-%m-%d"),
        "days": days,
        "planning_window_days": planning_window_days,
        "force_ship_window_days": force_ship_window_days,
        "daily_results": daily_results,
        "total_containers": total_containers,
        "total_shipped": len(shipped_ids),
        "total_pulls": total_pulls,
        "total_alert_containers": total_alert_containers,
        "alert_rate": round((total_alert_containers / total_containers) * 100, 1) if total_containers else 0,
        "weekly_container_rate": round((total_containers / days) * 7, 1) if days else 0,
        "remaining_count": len(remaining_items),
        "overdue_count": overdue_count,
        "deadline_remaining_count": overdue_count,
        "remaining_risks": build_remaining_risks(remaining_items, period_end),
        "avg_volume_rate": round(sum(volume_rates) / len(volume_rates), 1) if volume_rates else 0,
        "avg_weight_rate": round(sum(weight_rates) / len(weight_rates), 1) if weight_rates else 0,
    }

def rolling_planning_candidates(days, item_count=0):
    horizon = max(0, days)
    candidate_pool = ROLLING_PLANNING_WINDOW_CANDIDATES
    if item_count >= 750:
        candidate_pool = (3, 7, 14)
    elif item_count >= 300:
        candidate_pool = (0, 5, 7, 10, 14)
    candidates = [window for window in candidate_pool if window <= horizon]
    return candidates or [0]

def rolling_result_score(result):
    return (
        result["overdue_count"],
        result["total_alert_containers"],
        result["total_containers"],
        result["remaining_count"],
        -result["avg_volume_rate"],
        abs(result.get("planning_window_days", 0) - ROLLING_PLANNING_WINDOW_DAYS),
        result["total_pulls"],
    )

def summarize_rolling_trial(result):
    return {
        "planning_window_days": result.get("planning_window_days", 0),
        "force_ship_window_days": result.get("force_ship_window_days", 0),
        "total_containers": result["total_containers"],
        "total_shipped": result["total_shipped"],
        "total_pulls": result["total_pulls"],
        "total_alert_containers": result["total_alert_containers"],
        "alert_rate": result.get("alert_rate", 0),
        "weekly_container_rate": result.get("weekly_container_rate", 0),
        "remaining_count": result["remaining_count"],
        "overdue_count": result["overdue_count"],
        "avg_volume_rate": result["avg_volume_rate"],
        "avg_weight_rate": result["avg_weight_rate"],
        "score": rolling_result_score(result),
    }

def run_adaptive_rolling_simulation(items, start_date, days=30, allow_pull=True):
    trials = []
    best_result = None
    best_score = None

    for planning_window_days in rolling_planning_candidates(days, item_count=len(items)):
        result = run_rolling_simulation(
            items,
            start_date,
            days,
            allow_pull=allow_pull,
            planning_window_days=planning_window_days,
            force_ship_window_days=ROLLING_FORCE_SHIP_WINDOW_DAYS
        )
        score = rolling_result_score(result)
        trials.append(summarize_rolling_trial(result))
        if best_score is None or score < best_score:
            best_score = score
            best_result = result

    best_result["strategy"] = "adaptive_planning_window"
    best_result["strategy_trials"] = trials
    return best_result

def build_rolling_comparison(baseline, optimized):
    return {
        "baseline": {
            "total_containers": baseline["total_containers"],
            "total_shipped": baseline["total_shipped"],
            "total_pulls": baseline["total_pulls"],
            "total_alert_containers": baseline["total_alert_containers"],
            "alert_rate": baseline.get("alert_rate", 0),
            "weekly_container_rate": baseline.get("weekly_container_rate", 0),
            "remaining_count": baseline["remaining_count"],
            "overdue_count": baseline["overdue_count"],
            "avg_volume_rate": baseline["avg_volume_rate"],
            "avg_weight_rate": baseline["avg_weight_rate"],
        },
        "optimized": {
            "total_containers": optimized["total_containers"],
            "total_shipped": optimized["total_shipped"],
            "total_pulls": optimized["total_pulls"],
            "total_alert_containers": optimized["total_alert_containers"],
            "alert_rate": optimized.get("alert_rate", 0),
            "weekly_container_rate": optimized.get("weekly_container_rate", 0),
            "remaining_count": optimized["remaining_count"],
            "overdue_count": optimized["overdue_count"],
            "avg_volume_rate": optimized["avg_volume_rate"],
            "avg_weight_rate": optimized["avg_weight_rate"],
        },
        "delta": {
            "total_containers": baseline["total_containers"] - optimized["total_containers"],
            "total_shipped": optimized["total_shipped"] - baseline["total_shipped"],
            "total_alert_containers": baseline["total_alert_containers"] - optimized["total_alert_containers"],
            "alert_rate": round(baseline.get("alert_rate", 0) - optimized.get("alert_rate", 0), 1),
            "weekly_container_rate": round(baseline.get("weekly_container_rate", 0) - optimized.get("weekly_container_rate", 0), 1),
            "remaining_count": baseline["remaining_count"] - optimized["remaining_count"],
            "overdue_count": baseline["overdue_count"] - optimized["overdue_count"],
            "avg_volume_rate": round(optimized["avg_volume_rate"] - baseline["avg_volume_rate"], 1),
            "avg_weight_rate": round(optimized["avg_weight_rate"] - baseline["avg_weight_rate"], 1),
        }
    }

def run_single_day_scenario(items, current_date, must_ship_window_days):
    scenario_items = copy.deepcopy(items)
    reset_runtime_state(scenario_items)
    day_groups = classify_items_for_day(scenario_items, current_date, must_ship_window_days)
    engine = VanningEngine()
    containers, pool, unused_forwardable = engine.run_time_axis_packing(
        day_groups["must_ship"],
        day_groups["forwardable"],
        current_date,
        allow_partial_red_pull=True
    )
    container_summary = summarize_containers(containers)
    total_pulls = sum(
        1
        for container in containers
        for item in container.items
        if item.status_msg and "前倒し" in item.status_msg
    )
    return {
        "must_ship_window_days": must_ship_window_days,
        "container_count": container_summary["container_count"],
        "alert_containers": container_summary["alert_containers"],
        "avg_volume_rate": container_summary["avg_volume_rate"],
        "avg_weight_rate": container_summary["avg_weight_rate"],
        "total_pulls": total_pulls,
        "must_ship_count": len(day_groups["must_ship"]),
        "pool_count": len(pool),
        "unused_forwardable_count": len(unused_forwardable),
        "forwardable_count": len(day_groups["forwardable"]),
        "hold_count": len(day_groups["hold"]),
        "future_count": len(day_groups["future"]),
    }

def single_day_scenario_score(scenario):
    return (
        scenario["alert_containers"],
        scenario["container_count"],
        -scenario["avg_volume_rate"],
        scenario["pool_count"],
        scenario["unused_forwardable_count"],
        scenario["must_ship_window_days"],
    )

def load_case_master_rows():
    case_path = os.path.join(os.path.dirname(__file__), CASE_MASTER_FILE)
    if not os.path.exists(case_path):
        raise FileNotFoundError(f"ケースマスタが見つかりません: {CASE_MASTER_FILE}")

    case_mtime = os.path.getmtime(case_path)
    if (
        CASE_MASTER_CACHE["path"] == case_path
        and CASE_MASTER_CACHE["mtime"] == case_mtime
        and CASE_MASTER_CACHE["rows"] is not None
    ):
        return copy.deepcopy(CASE_MASTER_CACHE["rows"])

    df_raw = pd.read_excel(case_path, sheet_name=0, header=None)
    header_row_idx = 0
    for idx in range(min(10, len(df_raw))):
        values = {str(value).strip() for value in df_raw.iloc[idx].values if not pd.isna(value)}
        if {"分類", "資材名称", "L", "W", "H"}.issubset(values):
            header_row_idx = idx
            break

    df = pd.read_excel(case_path, sheet_name=0, header=header_row_idx)
    rows = []
    for _, row in df.iterrows():
        try:
            category = str(row["分類"]).strip()
            material_name = str(row["資材名称"]).strip()
            length = float(row["L"])
            width = float(row["W"])
            height = float(row["H"])
        except (KeyError, TypeError, ValueError):
            continue

        if not category or category == "nan" or not material_name or material_name == "nan":
            continue

        volume_m3 = (length * width * height) / 1_000_000_000
        footprint_m2 = (length * width) / 1_000_000
        rows.append({
            "分類": category,
            "No": row.get("№", row.get("No", "")),
            "資材名称": material_name,
            "L": int(length),
            "W": int(width),
            "H": int(height),
            "体積m3": round(volume_m3, 3),
            "床面積m2": round(footprint_m2, 3),
            "ケース階級": classify_case_profile(category, volume_m3, height),
        })

    if not rows:
        raise ValueError("ケースマスタから有効なケース寸法を読み取れませんでした。")
    CASE_MASTER_CACHE.update({"path": case_path, "mtime": case_mtime, "rows": copy.deepcopy(rows)})
    return rows

def classify_case_profile(category, volume_m3, height):
    if category == "木箱":
        if volume_m3 >= 8.0 or height >= 1800:
            return "大型木箱/特殊木箱"
        return "標準木箱"
    if volume_m3 >= 4.5 or height >= 1400:
        return "大型スチール"
    if volume_m3 >= 2.0:
        return "中型スチール"
    return "小型スチール"

def weighted_choice(randomizer, weighted_values):
    total = sum(weight for _, weight in weighted_values)
    cursor = randomizer.uniform(0, total)
    upto = 0
    for value, weight in weighted_values:
        upto += weight
        if cursor <= upto:
            return value
    return weighted_values[-1][0]

def generate_case_based_simulation_rows(case_master, row_count=200, assumptions=None, seed=20260516):
    import random

    assumptions = assumptions or {}
    randomizer = random.Random(seed)
    today = datetime.date.today()
    class_weights = assumptions["class_weights"]
    density_ranges = assumptions["density_ranges"]
    density_mix = assumptions["density_mix"]
    weight_variation = assumptions["weight_variation"]
    max_item_weight_kg = assumptions["max_item_weight_kg"]
    arrival_distribution = [
        (bucket, bucket.get("weight", 0))
        for bucket in assumptions["arrival_distribution"]
        if bucket.get("weight", 0) > 0
    ]

    by_class = {}
    for case in case_master:
        by_class.setdefault(case["ケース階級"], []).append(case)
    available_classes = [(case_class, class_weights.get(case_class, 0.1)) for case_class in by_class]

    rows = []
    for index in range(1, row_count + 1):
        case_class = weighted_choice(randomizer, available_classes)
        case = randomizer.choice(by_class[case_class])
        density_class = weighted_choice(randomizer, density_mix[case_class])
        density_min, density_max = density_ranges[density_class]
        density = randomizer.uniform(density_min, density_max)
        weight = max(40, min(max_item_weight_kg, case["体積m3"] * density * randomizer.uniform(*weight_variation)))
        weight = round(weight / 10) * 10

        arrival_bucket = weighted_choice(randomizer, arrival_distribution)
        creation_offset = randomizer.randint(*arrival_bucket["offset_days"])

        due_offset = max(
            creation_offset + randomizer.randint(*assumptions["due_after_arrival_days"]),
            randomizer.randint(*assumptions["due_offset_days"])
        )
        creation_date = today + datetime.timedelta(days=creation_offset)
        due_date = today + datetime.timedelta(days=due_offset)

        is_wood = case["分類"] == "木箱"
        expiration_date = ""
        if is_wood:
            expiration_offset = max(due_offset, creation_offset + randomizer.choice(assumptions["wood_expiration_days"]))
            expiration_date = (today + datetime.timedelta(days=expiration_offset)).strftime("%Y-%m-%d")

        stackable = "可"
        no_stack_probs = assumptions["stackable_unavailable_probability"]
        if case_class in no_stack_probs and randomizer.random() < no_stack_probs[case_class]:
            stackable = "不可"
        elif is_wood and randomizer.random() < no_stack_probs.get("木箱", 0):
            stackable = "不可"

        floor_only = "不可"
        floor_only_probs = assumptions["floor_only_probability"]
        if density_class == "重量物" and (weight >= 1200 or randomizer.random() < floor_only_probs.get("重量物", 0)):
            floor_only = "可"
        if case_class == "大型木箱/特殊木箱" and randomizer.random() < floor_only_probs.get("大型木箱/特殊木箱", 0):
            floor_only = "可"

        separation_group = ""
        if randomizer.random() < assumptions["separation_probability"]:
            separation_group = randomizer.choice(assumptions["separation_groups"])

        priority = next(
            row["priority"]
            for row in assumptions["priority_thresholds"]
            if due_offset <= row["due_days"]
        )
        allow_early_ship = "可" if randomizer.random() < assumptions["allow_early_ship_probability"] else "不可"
        item_name = f"ISZ-{case_class}-{case['資材名称']}-{index:04d}"

        rows.append({
            "品名": item_name,
            "ケース分類": case["分類"],
            "ケースNo": case["No"],
            "資材名称": case["資材名称"],
            "ケース階級": case_class,
            "L": case["L"],
            "W": case["W"],
            "H": case["H"],
            "体積m3": case["体積m3"],
            "重量": weight,
            "仮定重量階級": density_class,
            "積載可能日": creation_date.strftime("%Y-%m-%d"),
            "納期": due_date.strftime("%Y-%m-%d"),
            "木箱期限": expiration_date,
            "優先度": priority,
            "前倒し可否": allow_early_ship,
            "段積み可否": stackable,
            "床置き": floor_only,
            "分離グループ": separation_group,
            "仮定根拠": "寸法は開示ケースリスト準拠。重量・頻度・納期・制約は暫定仮定。",
        })
    return rows

def build_case_class_summary_rows(case_master):
    summary = {}
    for case in case_master:
        case_class = case["ケース階級"]
        bucket = summary.setdefault(case_class, {
            "ケース階級": case_class,
            "件数": 0,
            "分類": set(),
            "最小体積m3": None,
            "最大体積m3": None,
            "平均体積m3": 0,
            "最小L": None,
            "最大L": None,
            "最小W": None,
            "最大W": None,
            "最小H": None,
            "最大H": None,
        })
        bucket["件数"] += 1
        bucket["分類"].add(case["分類"])
        bucket["平均体積m3"] += case["体積m3"]
        for source, min_key, max_key in [
            ("体積m3", "最小体積m3", "最大体積m3"),
            ("L", "最小L", "最大L"),
            ("W", "最小W", "最大W"),
            ("H", "最小H", "最大H"),
        ]:
            value = case[source]
            bucket[min_key] = value if bucket[min_key] is None else min(bucket[min_key], value)
            bucket[max_key] = value if bucket[max_key] is None else max(bucket[max_key], value)

    rows = []
    class_order = ["小型スチール", "中型スチール", "大型スチール", "標準木箱", "大型木箱/特殊木箱"]
    for case_class in class_order:
        if case_class not in summary:
            continue
        bucket = summary[case_class]
        rows.append({
            "ケース階級": bucket["ケース階級"],
            "件数": bucket["件数"],
            "分類": "/".join(sorted(bucket["分類"])),
            "最小体積m3": round(bucket["最小体積m3"], 3),
            "最大体積m3": round(bucket["最大体積m3"], 3),
            "平均体積m3": round(bucket["平均体積m3"] / bucket["件数"], 3),
            "L範囲": f"{int(bucket['最小L'])} - {int(bucket['最大L'])}",
            "W範囲": f"{int(bucket['最小W'])} - {int(bucket['最大W'])}",
            "H範囲": f"{int(bucket['最小H'])} - {int(bucket['最大H'])}",
        })
    return rows

def build_case_assumptions_rows(row_count=200):
    return [
        {"項目": "ケース寸法", "扱い": "開示資料ベース", "内容": "いすゞロジスティック開示資料「ケースリスト」のL/W/Hをケースマスタとして使用"},
        {"項目": "ケース種類", "扱い": "開示資料ベース", "内容": "木箱10種類、スチール21種類を保持し、生成時は5つのケース階級へ分類"},
        {"項目": "ケース階級", "扱い": "モデル化", "内容": "小型スチール/中型スチール/大型スチール/標準木箱/大型木箱・特殊木箱"},
        {"項目": "運用規模", "扱い": "先方回答", "内容": f"週に50〜100本のバンニングレイアウトを行う規模感。今回の生成行数は{row_count}件"},
        {"項目": "重量", "扱い": "仮定", "内容": "軽量・標準・重量物の密度レンジをケース体積に掛けて生成。先方回答が来たら差し替え対象"},
        {"項目": "出現頻度", "扱い": "仮定", "内容": "ケース階級ごとの発生比率を暫定設定。実績頻度が分かり次第、置き換え可能"},
        {"項目": "日別入荷量/納期分布", "扱い": "仮定", "内容": "30日ローリング検証用に、到着済み・近日到着・将来到着を混在させて生成"},
        {"項目": "木箱期限", "扱い": "仮定", "内容": "木箱のみ、積載可能日から14/21/28日のいずれかで暫定付与し、納期より前には置かない"},
        {"項目": "段積み/床置き/分離制約", "扱い": "仮定", "内容": "大型・重量物ほど制約が出やすい前提。先方回答が来たら差し替え対象"},
        {"項目": "赤字判定", "扱い": "業務ルール", "内容": "体積充填率80%未満を赤字候補として評価。重量充填率は安全制約として扱う"},
    ]

@app.route('/api/assumptions', methods=['GET', 'POST'])
def assumptions_api():
    assumptions_path = os.path.join(os.path.dirname(__file__), SIMULATION_ASSUMPTIONS_FILE)
    assumptions = load_simulation_assumptions(assumptions_path)

    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        for key in ("allow_early_ship_probability", "separation_probability"):
            if key in data:
                try:
                    value = float(data[key])
                except (TypeError, ValueError):
                    return jsonify({'error': f'{key} は数値で指定してください'}), 400
                assumptions[key] = max(0.0, min(1.0, value))

        with open(assumptions_path, "w", encoding="utf-8") as handle:
            json.dump(assumptions, handle, ensure_ascii=False, indent=2)

    return jsonify({
        "allow_early_ship_probability": assumptions.get("allow_early_ship_probability", 0),
        "separation_probability": assumptions.get("separation_probability", 0),
        "class_weights": assumptions.get("class_weights", {}),
        "density_ranges": assumptions.get("density_ranges", {}),
        "arrival_distribution": assumptions.get("arrival_distribution", []),
    })

@app.route('/api/download_template', methods=['GET'])
def download_template():
    row_count = request.args.get("rows", default=200, type=int)
    if not row_count:
        row_count = 200
    row_count = max(50, min(row_count, 3000))
    seed = request.args.get("seed", default=20260516, type=int) or 20260516
    assumptions_path = os.path.join(os.path.dirname(__file__), SIMULATION_ASSUMPTIONS_FILE)
    assumptions = load_simulation_assumptions(assumptions_path)

    case_master = load_case_master_rows()
    simulation_rows = generate_case_based_simulation_rows(case_master, row_count=row_count, assumptions=assumptions, seed=seed)
    df = pd.DataFrame(simulation_rows)
    case_df = pd.DataFrame(case_master)
    summary_df = pd.DataFrame(build_case_class_summary_rows(case_master))
    parameters_df = pd.DataFrame(build_generation_parameter_rows(assumptions, row_count=row_count, seed=seed))
    assumptions_df = pd.DataFrame(build_case_assumptions_rows(row_count=row_count))
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='シミュレーションデータ')
        case_df.to_excel(writer, index=False, sheet_name='ケースマスタ')
        summary_df.to_excel(writer, index=False, sheet_name='ケース階級サマリ')
        parameters_df.to_excel(writer, index=False, sheet_name='生成パラメータ')
        assumptions_df.to_excel(writer, index=False, sheet_name='前提条件')

        for sheet_name in ["シミュレーションデータ", "ケースマスタ", "ケース階級サマリ", "生成パラメータ", "前提条件"]:
            ws = writer.book[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_length + 2, 10), 34)

        writer.book["生成パラメータ"].column_dimensions["E"].width = 56
        writer.book["前提条件"].column_dimensions["C"].width = 72
    output.seek(0)

    return send_file(
        output,
        download_name='case_master_based_simulation_data.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'ファイルがありません'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'ファイルが選択されていません'}), 400

    try:
        filepath = build_upload_path(file.filename)
        file.save(filepath)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    
    loader = DataLoader()
    try:
        items = loader.load_from_excel(filepath)
        validation_issues, invalid_item_ids = validate_items(items)
        clear_last_results()
        SESSION_DATA["items"] = items
        SESSION_DATA["validation_issues"] = validation_issues
        SESSION_DATA["invalid_item_ids"] = invalid_item_ids
        SESSION_DATA["input_profile"] = loader.input_profile
        generation_parameters = read_generation_parameters(filepath)
        simulation_context = build_simulation_context(items, loader.input_profile, generation_parameters)
        SESSION_DATA["simulation_context"] = simulation_context
        readiness = build_data_readiness(loader.input_profile, validation_issues)
        
        # サマリー作成（UI表示用）
        target_count = len(items[:20]) # テストとして最初の20件を今週分とする
        future_count = len(items[20:])
        
        return jsonify({
            'message': 'Success',
            'total_items': len(items),
            'valid_items': len(items) - len(invalid_item_ids),
            'invalid_items': len(invalid_item_ids),
            'target_count': target_count,
            'future_count': future_count,
            'validation_summary': build_validation_summary(validation_issues),
            'validation_issues': validation_issues[:8],
            'readiness': readiness,
            'simulation_context': simulation_context
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/optimize', methods=['POST'])
def optimize():
    data = request.get_json(silent=True) or {}
    base_date_str = data.get('base_date')
    must_ship_window_days = read_bounded_int(data, 'must_ship_window_days', 7, 0, 30)
    current_date, date_error = parse_base_date(base_date_str)
    if date_error:
        return jsonify({'error': date_error}), 400
        
    items = SESSION_DATA.get("items", [])
    if not items:
        return jsonify({'error': 'データがありません'}), 400

    valid_items, invalid_item_ids = build_valid_items()
    if not valid_items:
        return jsonify({'error': '積載可能な有効データがありません。入力データのエラーを確認してください。'}), 400

    reset_runtime_state(items)
    day_groups = classify_items_for_day(valid_items, current_date, must_ship_window_days)
    target_items = day_groups["must_ship"]
    forwardable_items = day_groups["forwardable"]
    hold_items = day_groups["hold"]
    future_items = day_groups["future"]

    baseline_engine = VanningEngine()
    baseline_containers, baseline_unpacked = baseline_engine.run_basic_packing(copy.deepcopy(target_items))
    
    engine = VanningEngine()
    containers, pool, unused_forwardable = engine.run_time_axis_packing(
        target_items,
        forwardable_items,
        current_date,
        allow_partial_red_pull=True
    )
    selected_packing_strategy = engine.last_packing_strategy
    containers = order_containers_for_output(containers)
    
    # 指示書ダウンロード用に保持
    SESSION_DATA["last_containers"] = containers
    SESSION_DATA["last_base_date"] = current_date
    SESSION_DATA["last_must_ship_window_days"] = must_ship_window_days
    SESSION_DATA["last_alert_summaries"] = {}
    
    # UI描画用に結果をJSON化
    result_containers = []
    for display_order, c in enumerate(containers, 1):
        is_red_ink = c.fill_rate_volume < VOLUME_TARGET_RATE
        c_items = []
        for i in c.items:
            c_items.append({
                'id': i.id,
                'name': i.name,
                'status_msg': i.status_msg,
                'selection_reason': i.selection_reason,
                'decision_reason': i.decision_reason or i.selection_reason,
                'constraint_tags': build_constraint_tags(i),
                'is_force_ship': i.is_force_ship,
                'is_manual_force_ship': i.force_ship,
                'is_system_force_ship': i.system_force_ship,
                'l': i.length,
                'w': i.width,
                'h': i.height,
                'x': i.x if i.x is not None else 0,
                'y': i.y if i.y is not None else 0,
                'z': i.z if i.z is not None else 0,
                'rotated': i.is_rotated
            })
            
        # 最短納期と最短木箱期限の計算
        earliest_due = min([i.due_date for i in c.items]) if c.items else None
        earliest_exp_items = [i.expiration_date for i in c.items if i.expiration_date]
        earliest_exp = min(earliest_exp_items) if earliest_exp_items else None
        
        pull_count = sum(1 for i in c.items if i.status_msg and "前倒し" in i.status_msg)
        blocker_counts = build_forwardable_blocker_counts(engine, c, unused_forwardable)
        alert_summary = build_container_alert_summary(c, current_date, must_ship_window_days, blocker_counts)
        SESSION_DATA["last_alert_summaries"][c.id] = alert_summary
        geometry_check = validate_container_geometry(c)
        
        cg_x, cg_y, cg_z = c.compute_cg()
            
        result_containers.append({
            'id': c.id,
            'display_order': display_order,
            'weight_val': c.current_weight,
            'weight_max': c.max_weight,
            'weight_rate': round(c.fill_rate_weight, 1),
            'volume_rate': round(c.fill_rate_volume, 1),
            'earliest_due': earliest_due.strftime('%Y-%m-%d') if earliest_due else 'なし',
            'earliest_exp': earliest_exp.strftime('%Y-%m-%d') if earliest_exp else '制限なし',
            'pull_count': pull_count,
            'is_alert': is_red_ink,
            'alert_reason_title': alert_summary["title"],
            'alert_reason_detail': alert_summary["detail"],
            'alert_blockers': blocker_counts,
            'geometry_valid': geometry_check["valid"],
            'geometry_warnings': geometry_check["warnings"],
            'geometry_warning_count': geometry_check["warning_count"],
            'cg_x': cg_x,
            'cg_y': cg_y,
            'cg_z': cg_z,
            'items': c_items
        })
        
    # 全体のサマリー集計
    total_pulls = sum(c['pull_count'] for c in result_containers)
    alert_containers = sum(1 for c in result_containers if c['is_alert'])
    comparison = build_comparison(baseline_containers, containers)
    validation_issues = SESSION_DATA.get("validation_issues", [])
    review_queue = build_review_queue(pool, unused_forwardable, hold_items, future_items, current_date)
    SESSION_DATA["last_review_queue"] = review_queue
        
    return jsonify({
        'containers': result_containers,
        'pool_count': len(pool),
        'baseline_unpacked_count': len(baseline_unpacked),
        'future_count': len(future_items),
        'forwardable_count': len(forwardable_items),
        'unused_forwardable_count': len(unused_forwardable),
        'hold_count': len(hold_items),
        'must_ship_count': len(target_items),
        'must_ship_window_days': must_ship_window_days,
        'review_queue': review_queue,
        'excluded_count': len(invalid_item_ids),
        'validation_summary': build_validation_summary(validation_issues),
        'validation_issues': validation_issues[:8],
        'readiness': build_data_readiness(SESSION_DATA.get("input_profile", {}), validation_issues),
        'comparison': comparison,
        'packing_strategy': selected_packing_strategy,
        'total_pulls': total_pulls,
        'alert_containers': alert_containers
    })

@app.route('/api/rolling', methods=['POST'])
def rolling_simulation():
    data = request.get_json(silent=True) or {}
    base_date_str = data.get('base_date')
    days = read_bounded_int(data, 'days', 30, 1, 90)

    current_date, date_error = parse_base_date(base_date_str)
    if date_error:
        return jsonify({'error': date_error}), 400

    valid_items, invalid_item_ids = build_valid_items()
    if not valid_items:
        return jsonify({'error': '積載可能な有効データがありません。入力データのエラーを確認してください。'}), 400

    baseline = run_rolling_simulation(
        valid_items,
        current_date,
        days,
        allow_pull=False,
        planning_window_days=0,
        force_ship_window_days=ROLLING_FORCE_SHIP_WINDOW_DAYS
    )
    result = run_adaptive_rolling_simulation(
        valid_items,
        current_date,
        days,
        allow_pull=True
    )
    result.update({
        'excluded_count': len(invalid_item_ids),
        'validation_summary': build_validation_summary(SESSION_DATA.get("validation_issues", [])),
        'validation_issues': SESSION_DATA.get("validation_issues", [])[:8],
        'readiness': build_data_readiness(SESSION_DATA.get("input_profile", {}), SESSION_DATA.get("validation_issues", [])),
        'comparison': build_rolling_comparison(baseline, result),
    })
    SESSION_DATA["last_rolling"] = result
    return jsonify(result)

@app.route('/api/scenarios', methods=['POST'])
def compare_scenarios():
    data = request.get_json(silent=True) or {}
    base_date_str = data.get('base_date')
    current_window = read_bounded_int(data, 'must_ship_window_days', 7, 0, 30)
    current_date, date_error = parse_base_date(base_date_str)
    if date_error:
        return jsonify({'error': date_error}), 400

    valid_items, invalid_item_ids = build_valid_items()
    if not valid_items:
        return jsonify({'error': '積載可能な有効データがありません。入力データのエラーを確認してください。'}), 400

    requested_windows = data.get("windows")
    if isinstance(requested_windows, list):
        windows = sorted({max(0, min(30, int(window))) for window in requested_windows if str(window).lstrip("-").isdigit()})
    else:
        windows = sorted({3, 7, 10, 14, current_window})
    if not windows:
        windows = [current_window]

    scenarios = [
        run_single_day_scenario(valid_items, current_date, window)
        for window in windows
    ]
    best = min(scenarios, key=single_day_scenario_score)
    for scenario in scenarios:
        scenario["recommended"] = scenario["must_ship_window_days"] == best["must_ship_window_days"]

    SESSION_DATA["last_scenarios"] = {
        "base_date": current_date.strftime("%Y-%m-%d"),
        "scenarios": scenarios,
        "recommended": best,
        "excluded_count": len(invalid_item_ids),
    }

    return jsonify({
        "base_date": current_date.strftime("%Y-%m-%d"),
        "scenarios": scenarios,
        "recommended": best,
        "excluded_count": len(invalid_item_ids),
        "validation_summary": build_validation_summary(SESSION_DATA.get("validation_issues", [])),
        "readiness": build_data_readiness(SESSION_DATA.get("input_profile", {}), SESSION_DATA.get("validation_issues", [])),
    })

@app.route('/api/export_rolling', methods=['GET'])
def export_rolling_excel():
    rolling = SESSION_DATA.get("last_rolling")
    if not rolling:
        return jsonify({'error': 'エクスポートするローリング結果がありません'}), 400

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output = BytesIO()
    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    summary_ws.merge_cells('A1:D1')
    summary_ws['A1'] = f"【30日ローリングシミュレーション】 {rolling['start_date']} 〜 {rolling['end_date']}"
    summary_ws['A1'] = f"【{rolling['days']}日ローリングシミュレーション】 {rolling['start_date']} 〜 {rolling['end_date']}"
    summary_ws['A1'].font = Font(size=14, bold=True)

    summary_rows = [
        ("計画先読み日数", rolling.get("planning_window_days", 0), "日"),
        ("赤字強制出荷判定日数", rolling.get("force_ship_window_days", 0), "日"),
        ("期間日数", rolling["days"], "日"),
        ("総出荷数", rolling["total_shipped"], "件"),
        ("総コンテナ数", rolling["total_containers"], "本"),
        ("週換算コンテナ数", rolling.get("weekly_container_rate", 0), "本/週"),
        ("前倒し補填数", rolling["total_pulls"], "件"),
        ("赤字コンテナ数", rolling["total_alert_containers"], "本"),
        ("赤字率", rolling.get("alert_rate", 0), "%"),
        ("平均体積充填率", rolling["avg_volume_rate"], "%"),
        ("平均重量充填率", rolling["avg_weight_rate"], "%"),
        ("期間後残", rolling["remaining_count"], "件"),
        ("期限到来/超過残", rolling["overdue_count"], "件"),
        ("入力除外", rolling.get("excluded_count", 0), "件"),
    ]

    for row_idx, row in enumerate(summary_rows, 3):
        for col_idx, value in enumerate(row, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin
            if col_idx == 1:
                cell.fill = note_fill
                cell.font = Font(bold=True)

    comparison = rolling.get("comparison")
    if comparison:
        start_row = 16
        summary_ws[f'A{start_row}'] = "前倒しなし / あり 比較"
        summary_ws[f'A{start_row}'].font = Font(bold=True)
        headers = ["KPI", "前倒しなし", "前倒しあり", "差分"]
        for col_idx, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=start_row + 1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_thin

        comparison_rows = [
            ("総コンテナ数", "total_containers", "本"),
            ("総出荷数", "total_shipped", "件"),
            ("赤字コンテナ数", "total_alert_containers", "本"),
            ("期間後残", "remaining_count", "件"),
            ("期限到来/超過残", "overdue_count", "件"),
            ("平均体積充填率", "avg_volume_rate", "pt"),
            ("平均重量充填率", "avg_weight_rate", "pt"),
        ]
        for offset, (label, key, unit) in enumerate(comparison_rows, 2):
            row_idx = start_row + offset
            baseline = comparison["baseline"][key]
            optimized = comparison["optimized"][key]
            delta = comparison["delta"][key]
            values = [label, baseline, optimized, f"{delta:+g} {unit}"]
            for col_idx, value in enumerate(values, 1):
                cell = summary_ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border_thin

    daily_ws = wb.create_sheet(title="Daily")
    daily_headers = [
        "日付", "コンテナ数", "出荷件数", "前倒し補填", "赤字コンテナ",
        "必須出荷", "前倒し候補", "未使用前倒し候補", "前倒し不可",
        "未到着", "保留プール", "期限到来/超過残", "平均体積充填率", "平均重量充填率"
    ]
    for col_idx, header in enumerate(daily_headers, 1):
        cell = daily_ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin

    for row_idx, day in enumerate(rolling["daily_results"], 2):
        values = [
            day["date"], day["containers"], day["shipped"], day["pulls"], day["alerts"],
            day["must_ship"], day["forwardable"], day["unused_forwardable"], day["hold"],
            day["future"], day["pool"], day["overdue_remaining"], day["avg_volume_rate"], day["avg_weight_rate"]
        ]
        for col_idx, value in enumerate(values, 1):
            cell = daily_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin

    risk_ws = wb.create_sheet(title="RemainingRisks")
    risk_headers = [
        "品名", "確認理由", "積載可能日", "納期", "木箱期限", "納期まで(日)",
        "木箱期限まで(日)", "優先度", "前倒し可否", "実務制約", "分類理由", "判断根拠", "ステータス"
    ]
    for col_idx, header in enumerate(risk_headers, 1):
        cell = risk_ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin

    for row_idx, item in enumerate(rolling.get("remaining_risks", []), 2):
        values = [
            item["name"],
            item["reason"],
            item["creation_date"],
            item["due_date"],
            item["expiration_date"],
            item["days_until_due"],
            item["days_until_expiration"],
            item["priority"],
            "可" if item["allow_early_ship"] else "不可",
            " / ".join(item["constraint_tags"]),
            item["selection_reason"],
            item["decision_reason"],
            item["status_msg"],
        ]
        for col_idx, value in enumerate(values, 1):
            cell = risk_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin

    trials = rolling.get("strategy_trials", [])
    trial_ws = None
    if trials:
        trial_ws = wb.create_sheet(title="StrategyTrials")
        trial_headers = [
            "計画先読み日数", "赤字強制出荷判定日数", "コンテナ数", "出荷件数", "前倒し補填",
            "赤字コンテナ", "赤字率", "週換算コンテナ数", "期間後残", "期限到来/超過残",
            "平均体積充填率", "平均重量充填率", "採用"
        ]
        for col_idx, header in enumerate(trial_headers, 1):
            cell = trial_ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin

        selected_window = rolling.get("planning_window_days")
        selected_force_window = rolling.get("force_ship_window_days")
        for row_idx, trial in enumerate(trials, 2):
            is_selected = (
                trial.get("planning_window_days") == selected_window
                and trial.get("force_ship_window_days") == selected_force_window
            )
            values = [
                trial.get("planning_window_days", 0),
                trial.get("force_ship_window_days", 0),
                trial.get("total_containers", 0),
                trial.get("total_shipped", 0),
                trial.get("total_pulls", 0),
                trial.get("total_alert_containers", 0),
                trial.get("alert_rate", 0),
                trial.get("weekly_container_rate", 0),
                trial.get("remaining_count", 0),
                trial.get("overdue_count", 0),
                trial.get("avg_volume_rate", 0),
                trial.get("avg_weight_rate", 0),
                "採用" if is_selected else "",
            ]
            for col_idx, value in enumerate(values, 1):
                cell = trial_ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border_thin
                if is_selected:
                    cell.fill = note_fill

    export_sheets = [summary_ws, daily_ws, risk_ws]
    if trial_ws:
        export_sheets.append(trial_ws)

    for ws in export_sheets:
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name='rolling_simulation_report.xlsx',
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/api/export_scenarios', methods=['GET'])
def export_scenarios_excel():
    scenario_result = SESSION_DATA.get("last_scenarios")
    if not scenario_result:
        return jsonify({'error': 'エクスポートするシナリオ比較結果がありません'}), 400

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ScenarioComparison"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:L1')
    ws['A1'] = f"【シナリオ比較】 基準日 {scenario_result['base_date']}"
    ws['A1'].font = Font(size=14, bold=True)

    recommended = scenario_result.get("recommended", {})
    ws['A3'] = "推奨シナリオ"
    ws['B3'] = f"必須出荷幅 {recommended.get('must_ship_window_days', '-')}日"
    ws['C3'] = f"赤字 {recommended.get('alert_containers', '-')}本"
    ws['D3'] = f"平均体積 {recommended.get('avg_volume_rate', '-')}%"
    for cell in ws[3]:
        cell.border = border_thin
        if cell.column == 1:
            cell.fill = note_fill
            cell.font = Font(bold=True)

    headers = [
        "推奨", "必須出荷幅(日)", "コンテナ数", "赤字コンテナ", "平均体積充填率",
        "平均重量充填率", "必須出荷件数", "前倒し補填", "保留プール",
        "前倒し候補残", "前倒し不可", "未到着"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin

    for row_idx, scenario in enumerate(scenario_result.get("scenarios", []), 6):
        values = [
            "採用" if scenario.get("recommended") else "",
            scenario.get("must_ship_window_days"),
            scenario.get("container_count"),
            scenario.get("alert_containers"),
            scenario.get("avg_volume_rate"),
            scenario.get("avg_weight_rate"),
            scenario.get("must_ship_count"),
            scenario.get("total_pulls"),
            scenario.get("pool_count"),
            scenario.get("unused_forwardable_count"),
            scenario.get("hold_count"),
            scenario.get("future_count"),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin
            if scenario.get("recommended"):
                cell.fill = note_fill

    widths = [10, 16, 12, 14, 16, 16, 14, 12, 12, 14, 12, 12]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        download_name='scenario_comparison_report.xlsx',
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/api/export', methods=['GET'])
def export_excel():
    containers = SESSION_DATA.get("last_containers", [])
    if not containers:
        return jsonify({'error': 'エクスポートするデータがありません'}), 400
        
    output = BytesIO()
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # デフォルトシートを削除
    
    # スタイルの定義
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    export_base_date = SESSION_DATA.get("last_base_date", datetime.date.today())
    export_must_ship_window_days = SESSION_DATA.get("last_must_ship_window_days", 7)
    alert_summaries = SESSION_DATA.get("last_alert_summaries", {})

    summary_ws = wb.create_sheet(title="納品順サマリー")
    summary_ws.merge_cells('A1:I1')
    summary_ws['A1'] = "【納品順サマリー】 コンテナ別出荷判断一覧"
    summary_ws['A1'].font = Font(size=14, bold=True)
    summary_headers = [
        "納品順", "計画ID", "最短納期", "木箱期限", "体積充填率", "重量充填率",
        "荷物数", "前倒し補填", "出荷判断"
    ]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin

    for row_idx, c in enumerate(containers, 4):
        earliest_due = min([item.due_date for item in c.items]) if c.items else None
        exp_items = [item.expiration_date for item in c.items if item.expiration_date]
        earliest_exp = min(exp_items) if exp_items else None
        pull_count = sum(1 for item in c.items if item.status_msg and "前倒し" in item.status_msg)
        alert_summary = alert_summaries.get(c.id) or build_container_alert_summary(c, export_base_date, export_must_ship_window_days)
        decision_text = (
            f"{alert_summary['title']}：{alert_summary['detail']}"
            if alert_summary.get("title")
            else "体積80%クリア"
        )
        values = [
            row_idx - 3,
            c.id,
            earliest_due.strftime("%Y-%m-%d") if earliest_due else "",
            earliest_exp.strftime("%Y-%m-%d") if earliest_exp else "制限なし",
            f"{round(c.fill_rate_volume, 1)}%",
            f"{round(c.fill_rate_weight, 1)}%",
            len(c.items),
            pull_count,
            decision_text,
        ]
        for col_idx, value in enumerate(values, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin
            if col_idx in {1, 7, 8}:
                cell.alignment = Alignment(horizontal="right")
            if c.fill_rate_volume < VOLUME_TARGET_RATE and col_idx == 9:
                cell.font = Font(color="FF0000", bold=True)
            if col_idx == 9:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    summary_widths = [10, 12, 14, 14, 14, 14, 10, 12, 72]
    for col_idx, width in enumerate(summary_widths, 1):
        summary_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    simulation_context = SESSION_DATA.get("simulation_context")
    if simulation_context:
        def pct(value):
            return "" if value is None else f"{value}%"

        context_ws = wb.create_sheet(title="前提サマリー")
        context_ws.merge_cells('A1:D1')
        context_ws['A1'] = "【前提サマリー】 今回のデータの根拠と仮定"
        context_ws['A1'].font = Font(size=14, bold=True)
        context_ws['A2'] = simulation_context.get("summary_note", "")
        context_ws.merge_cells('A2:D2')
        context_ws['A2'].alignment = Alignment(wrap_text=True, vertical="top")

        context_rows = [
            ("データ種別", "開示ケースリスト準拠の検証データ" if simulation_context.get("is_generated") else "読み込みExcelデータ"),
            ("読み込み荷物数", simulation_context.get("loaded_count", "")),
            ("生成時の荷物数", simulation_context.get("generated_rows_setting", "")),
            ("再現用番号", simulation_context.get("seed", "")),
            ("早めに積める荷物", f"設定 {pct(simulation_context.get('configured_early_ship_rate'))} / 実データ {pct(simulation_context.get('early_ship_rate'))}"),
            ("一緒に積めない指定", f"設定 {pct(simulation_context.get('configured_separation_rate'))} / 実データ {pct(simulation_context.get('separation_rate'))}"),
            ("段積み不可", pct(simulation_context.get("no_stack_rate"))),
            ("床置き指定", pct(simulation_context.get("floor_only_rate"))),
            ("木箱期限あり", pct(simulation_context.get("wood_deadline_rate"))),
            ("到着状況", f"到着済み {pct(simulation_context.get('arrived_rate'))} / 未到着 {pct(simulation_context.get('future_rate'))}"),
            ("納期範囲", f"{simulation_context.get('due_min', '')} 〜 {simulation_context.get('due_max', '')}"),
            ("平均重量", simulation_context.get("avg_weight", "")),
        ]
        for row_idx, (label, value) in enumerate(context_rows, 4):
            context_ws.cell(row=row_idx, column=1).value = label
            context_ws.cell(row=row_idx, column=2).value = value
            for col_idx in (1, 2):
                cell = context_ws.cell(row=row_idx, column=col_idx)
                cell.border = border_thin
                if col_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        source_start = 18
        source_headers = ["項目", "扱い", "根拠", "備考"]
        for col_idx, header in enumerate(source_headers, 1):
            cell = context_ws.cell(row=source_start, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin
        for row_idx, row in enumerate(simulation_context.get("source_rows", []), source_start + 1):
            values = [row.get("item", ""), row.get("status", ""), row.get("basis", ""), row.get("note", "")]
            for col_idx, value in enumerate(values, 1):
                cell = context_ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border_thin
                if row.get("status") in {"仮定", "補完"} and col_idx == 2:
                    cell.font = Font(color="B45309", bold=True)
                if col_idx == 4:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, width in enumerate([24, 18, 24, 72], 1):
            context_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    readiness = build_data_readiness(SESSION_DATA.get("input_profile", {}), SESSION_DATA.get("validation_issues", []))
    readiness_ws = wb.create_sheet(title="実務準備度")
    readiness_ws.merge_cells('A1:E1')
    readiness_ws['A1'] = f"【実務準備度】 {readiness['score']}% / {readiness['risk_level']}"
    readiness_ws['A1'].font = Font(size=14, bold=True)
    readiness_headers = ["項目", "状態", "確認列", "備考", "実データ"]
    for col_idx, header in enumerate(readiness_headers, 1):
        cell = readiness_ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_thin
    for row_idx, row in enumerate(readiness["checks"], 4):
        values = [row["item"], row["status"], row["source"], row["note"], "あり" if row["ready"] else "補完"]
        for col_idx, value in enumerate(values, 1):
            cell = readiness_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border_thin
            if not row["ready"] and col_idx in {2, 5}:
                cell.font = Font(color="FF0000", bold=True)
            if col_idx == 4:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx, width in enumerate([18, 14, 24, 72, 12], 1):
        readiness_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    for c in containers:
        ws = wb.create_sheet(title=str(c.id))
        alert_summary = alert_summaries.get(c.id) or build_container_alert_summary(c, export_base_date, export_must_ship_window_days)
        alert_text = (
            f"{alert_summary['title']}：{alert_summary['detail']}"
            if alert_summary["title"]
            else "体積80%クリア"
        )
        
        # コンテナサマリー部分
        ws.merge_cells('A1:I1')
        ws['A1'] = f"【バンニング指示書】 コンテナ: {c.id}"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = "最大重量(kg)"
        ws['B3'] = c.max_weight
        ws['A4'] = "現在重量(kg)"
        ws['B4'] = c.current_weight
        ws['A5'] = "重量充填率"
        ws['B5'] = f"{round(c.fill_rate_weight, 1)}%"
        ws['A6'] = "体積充填率"
        ws['B6'] = f"{round(c.fill_rate_volume, 1)}%"
        ws['A7'] = "出荷判断"
        ws['B7'] = alert_text
        ws.merge_cells('B7:I7')
        
        for r in range(3, 8):
            ws[f'A{r}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws[f'A{r}'].border = border_thin
            ws[f'B{r}'].border = border_thin
        ws['B7'].alignment = Alignment(wrap_text=True, vertical="top")
        if c.fill_rate_volume < VOLUME_TARGET_RATE:
            ws['B7'].font = Font(color="FF0000", bold=True)
        
        # 荷物リストのヘッダー
        headers = ["積込順", "品名", "重量(kg)", "寸法 L×W×H(mm)", "3D配置 (奥×横×高)", "実務制約", "分類理由", "判断根拠", "特記事項・ステータス"]
        start_row = 10
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin
            
        # 荷物データの書き込み
        for idx, i in enumerate(c.items, 1):
            row_idx = start_row + idx
            
            # 回転考慮の寸法
            dim_str = f"{i.length} × {i.width} × {i.height}" + (" (回転)" if i.is_rotated else "")
            
            # 3D座標 (Three.jsの表示に合わせて見やすく)
            pos_str = f"X:{i.x} Y:{i.y} Z:{i.z}" if i.x is not None else "未定"
            
            # 現場ステータス
            status = i.status_msg if i.status_msg else "通常"
            if i.is_force_ship and status == "通常":
                status = "🚨 赤字/納期強制出荷"
                
            row_data = [
                idx,
                i.name,
                f"{i.weight:,.1f}",
                dim_str,
                pos_str,
                " / ".join(build_constraint_tags(i)),
                i.selection_reason if i.selection_reason else "通常",
                i.decision_reason or i.selection_reason,
                status
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.border = border_thin
                if col_idx == 1 or col_idx == 3:
                    cell.alignment = Alignment(horizontal="right")
                    
                # 色付けルールの適用
                if col_idx == 9:
                    if "強制出荷" in status:
                        cell.font = Font(color="FF0000", bold=True) # 赤
                    elif "前倒し" in status:
                        cell.font = Font(color="008000", bold=True) # 緑
        
        # 列幅の自動調整
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 48
        ws.column_dimensions['I'].width = 40

    review_queue = SESSION_DATA.get("last_review_queue", [])
    if review_queue:
        ws = wb.create_sheet(title="ReviewQueue")
        headers = [
            "品名", "確認理由", "積載可能日", "納期", "木箱期限", "納期まで(日)",
            "木箱期限まで(日)", "優先度", "前倒し可否", "実務制約", "分類理由", "判断根拠", "ステータス"
        ]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin

        for row_idx, item in enumerate(review_queue, 2):
            row_data = [
                item["name"],
                item["reason"],
                item["creation_date"],
                item["due_date"],
                item["expiration_date"],
                item["days_until_due"],
                item["days_until_expiration"],
                item["priority"],
                "可" if item["allow_early_ship"] else "不可",
                " / ".join(item["constraint_tags"]),
                item["selection_reason"],
                item["decision_reason"],
                item["status_msg"],
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.border = border_thin

        widths = [30, 36, 14, 14, 14, 14, 16, 10, 12, 22, 20, 48, 30]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(output)
    output.seek(0)
    return send_file(
        output, 
        download_name='vanning_instructions.xlsx', 
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/api/override', methods=['POST'])
def override_item():
    data = request.get_json(silent=True) or {}
    item_ids = data.get('item_ids', [])
    force_ship = data.get('force_ship')
    if not isinstance(item_ids, list) or not item_ids:
        return jsonify({'error': '対象荷物が指定されていません'}), 400
    if not isinstance(force_ship, bool):
        return jsonify({'error': 'force_ship は true / false で指定してください'}), 400
    
    items = SESSION_DATA.get("items", [])
    found = False
    target_ids = set(str(item_id) for item_id in item_ids)
    for item in items:
        if item.id in target_ids:
            item.force_ship = force_ship
            item.status_msg = "手動オーバーライド：現場指示による強制出荷" if force_ship else ""
            found = True
            
    if found:
        clear_last_results()
        return jsonify({'message': 'Success'})
    return jsonify({'error': 'アイテムが見つかりません'}), 404

if __name__ == '__main__':
    print("サーバーを起動しました。ブラウザで http://127.0.0.1:5000 にアクセスしてください。")
    app.run(debug=True, port=5000)
